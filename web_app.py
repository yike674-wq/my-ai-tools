import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openai import OpenAI
import plotly.express as px
from io import BytesIO
from datetime import datetime

# --- 1. 稳健的初始化配置 ---
st.set_page_config(page_title="AI 审计终端 V12.3", page_icon="🧼", layout="wide")

# 确保所有必要的键都在 session_state 中，且初始值类型正确
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

# 重点修复：显式初始化为列表，防止 TypeError
if "messages" not in st.session_state or st.session_state["messages"] is None:
    st.session_state["messages"] = []

# 其他变量初始化
for key in ["df_cleaned", "current_file", "raw_binary"]:
    if key not in st.session_state:
        st.session_state[key] = None

# --- 2. 视觉引擎：颜色识别与提取 ---
def process_visual_data(file_bytes, mode="all"):
    """
    mode "uncolored": 仅提取没颜色的
    mode "colored": 仅提取标记了颜色的
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    extracted_data = []
    
    for row in ws.iter_rows(min_row=1):
        cell = row[0] # 默认检查第一列 A列
        fill = cell.fill
        # 识别颜色逻辑：'00000000' 或索引 64/0 通常代表无填充
        is_colored = fill.start_color.index not in ['00000000', 0, 64] and fill.fill_type is not None
        
        val = cell.value
        if val is not None:
            if mode == "uncolored" and not is_colored:
                extracted_data.append(val)
            elif mode == "colored" and is_colored:
                extracted_data.append(val)
            elif mode == "all":
                extracted_data.append(val)
                
    # 转换为 DataFrame 方便展示，由于是纯数字，自动命名为“号码库”
    return pd.DataFrame(extracted_data, columns=["号码库"])

# --- 3. 物理清洗引擎：生成干净的 Excel ---
def export_cleaned_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=True)
    return output.getvalue()

# --- 4. 侧边栏：清洗控制台 ---
with st.sidebar:
    st.title("🧼 数据清洗站")
    uploaded_file = st.file_uploader("📂 上传您的表格 (xlsx)", type=["xlsx"])
    
    if uploaded_file:
        st.session_state["raw_binary"] = uploaded_file.read()
        st.session_state["current_file"] = uploaded_file.name
        
        st.divider()
        st.subheader("🛠️ 视觉过滤选项")
        filter_mode = st.radio("提取范围：", ["全部提取", "仅提取未标记颜色", "仅提取已标记颜色"])
        
        mode_map = {"全部提取": "all", "仅提取未标记颜色": "uncolored", "仅提取已标记颜色": "colored"}
        
        if st.button("🚀 执行视觉提取", use_container_width=True):
            with st.spinner("正在扫描单元格颜色..."):
                st.session_state["df_cleaned"] = process_visual_data(
                    st.session_state["raw_binary"], 
                    mode=mode_map[filter_mode]
                )
            st.success("提取完成！")

    if st.session_state["df_cleaned"] is not None:
        st.divider()
        st.subheader("📥 成果导出")
        # 导出清洗后的 Excel
        clean_xlsx = export_cleaned_excel(st.session_state["df_cleaned"])
        st.download_button(
            label="💾 下载清洗后的 Excel",
            data=clean_xlsx,
            file_name=f"已清洗_{datetime.now().strftime('%H%m%s')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# --- 5. 主界面：审计与对话 ---
st.title("📊 AI 自动化办公看板 V12.3")

if st.session_state["df_cleaned"] is not None:
    df = st.session_state["df_cleaned"]
    
    tab_data, tab_ai = st.tabs(["💎 数据预览", "🤖 AI 深度诊断"])
    
    with tab_data:
        st.write(f"已为您提取 {len(df)} 条记录")
        st.dataframe(df, use_container_width=True)
        # 简单统计
        if not df.empty:
            st.info(f"💡 发现 {df.duplicated().sum()} 条重复记录")

    with tab_ai:
        st.subheader("🤖 首席 AI 审计官")
        
        # 1. 强制显示当前消息
        if not st.session_state["messages"]:
            st.info("💡 暂无对话记录。您可以尝试问：'这些号码中有重复的吗？'")
        else:
            for msg in st.session_state["messages"]:
                with st.chat_message(msg["role"]):
                    st.write(msg["content"])
        
        # 2. 检查 API Key 并显化输入框
        api_key = st.secrets.get("DEEPSEEK_API_KEY")
        
        if not api_key:
            st.error("🔑 未检测到 API Key！请在 Streamlit Secrets 中配置 DEEPSEEK_API_KEY。")
            # 即使没 Key，也给个模拟输入框样式，方便调试
            st.text_input("对话框已禁用 (缺少 Key)", disabled=True, placeholder="请先配置 API Key...")
        else:
            # 核心对话输入
            if prompt := st.chat_input("说点什么..."):
                st.session_state.messages.append({"role": "user", "content": prompt})
                with st.chat_message("user"):
                    st.write(prompt)
                
                with st.chat_message("assistant"):
                    client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
                    context = f"数据预览：\n{df.head(20).to_string()}"
                    response = st.write_stream(client.chat.completions.create(
                        model="deepseek-chat",
                        messages=[
                            {"role": "system", "content": f"审计专家。{context}"},
                            {"role": "user", "content": prompt}
                        ],
                        stream=True
                    ))
                st.session_state.messages.append({"role": "assistant", "content": response})
                st.rerun() # 强制刷新以保持对话框在底部

        if st.button("🗑️ 清空当前对话"):
            st.session_state["messages"] = []
            st.rerun()